import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# from pathlib import Path

# filep = Path('transactions.xlsx')
# print(filep.exists())

# wb = xl.load_workbook('transactions.xlsx')
# sheet = wb['Sheet1']
# cell = sheet['a1']
# sheet.cell(1, 1)


filename = 'transactions.xlsx'
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        if isinstance(cell.value, (int, float)):  # Ensure the value is numeric
            corrected_price = cell.value * 0.9
        else:
            corrected_price = 0  # Default to 0 if the value is invalid
    
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row, 
            min_col=4,
            max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save(filename)

process_workbook(filename)

print(xl.load_workbook(filename)['Sheet1'].cell(3, 4).value)
